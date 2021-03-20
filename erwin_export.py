import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Specifying the ODBC driver, server name, database, etc. directly
cnxn = pyodbc.connect('DSN=erwin_r9_current',server='127.0.0.1')
# Create a cursor from the connection
cursor = cnxn.cursor()

def fetch_model_name():
    sql = """
    SELECT 
    TRAN(PEn.owner_path) "Model Name"
    FROM M0.Entity Pen
    """

    cursor.execute(sql)
    model_name = cursor.fetchone()[0]

    return model_name

def create_col_headings(ws):
    # Grab the column names
    columns = [column[0] for column in cursor.description]
    col_ctr = 1
    for col in columns:
        c = ws.cell(row=1, column=col_ctr)
        c.font=Font(bold=True)
        c.value = col
        col_ctr = col_ctr + 1

def extract_defs(type_name, sql):
    cursor.execute(sql)

    # Create a worksheet if we're dealing with subject areas, entities or relationships
    #if type_name in ("Subject Areas", "Entities","Relationships"):
    ws = wb.create_sheet(type_name)
    max_col_widths = create_col_headings(ws)

    # Fetch and display the rows
    resultset = cursor.fetchall()

    #Initialise entity_id to a non-existent value
    entity_id = -1

    row_ctr = 2

    for row in resultset:
        col_ctr = 1
        for col in row:
            #Overwrite any attribute definition until DG provides
            if type_name == "Attributes" and col_ctr==5:
                col="To be defined by data governance"

            #Store the column value
            c = ws.cell(row=row_ctr, column=col_ctr)
            c.value = col

            col_ctr = col_ctr + 1
        row_ctr = row_ctr + 1


# Create an Excel workbook object
wb = Workbook()

#fetch all of the subject areas
sql="""
SELECT
   TRAN(SA.Long_Id) "Subject Area ID",
   SA.NAME                      'Subject Area',
   Tran(SA.Definition)          'Subject Area Definition' 
FROM
    SUBJECT_AREA SA
WHERE SA.Name is not null  
order By 1
"""

extract_defs("Subject Areas",sql)

# fetch all of the entities
sql = """
SELECT distinct
   TRAN(PEn.Long_Id) "Entity ID",
   SA.NAME                      'Subject Area',
   replace(Tran(PEn.Physical_Name),'_',' ')      'Entity Name',
   Tran(PEn.Definition)         'Entity Definition' 
FROM
   Entity PEn
      LEFT JOIN ER_MODEL_SHAPE EMS
        ON EMS.MODEL_OBJECT_REF = PEn.ID@
      LEFT JOIN ER_DIAGRAM ED
        ON ED.ID@ = EMS.OWNER@
      LEFT JOIN SUBJECT_AREA SA
        ON SA.ID@ = ED.OWNER@
WHERE SA.Name is not null  
and PEn.hide_in_logical is null
order By 2,3
            """
extract_defs("Entities",sql)

sql="""
SELECT distinct
TRAN(R.Long_Id) "Relationship ID",
P.NAME AS 'PARENT',
R.PARENT_TO_CHILD_VERB_PHRASE AS 'RELATIONSHIP',
C.NAME AS CHILD
FROM M0.RELATIONSHIP R INNER JOIN M0.ENTITY P
ON R.PARENT_ENTITY_REF = P.ID@
INNER JOIN M0.ENTITY C
ON R.CHILD_ENTITY_REF = C.ID@
WHERE C.HIDE_IN_LOGICAL is null
AND P.HIDE_IN_LOGICAL is null
ORDER by 2,4
"""
extract_defs("Relationships", sql)

# fetch all of the attributes
sql = """
SELECT distinct
TRAN(PEn.Long_Id) "Entity ID",
TRAN(PEn.Name) "Entity Name",
TRAN(PAt.Long_Id) "Attribute ID",
TRAN(PAt.Name) "Attribute Name",
TRAN(PAt.Definition) "Attribute Definition",
TRAN(PAt.Logical_Data_Type) "Attribute Data Type",
TRAN(PAt.Type) "Primary Key Flag",
TRAN(PAt.Null_Option_Type) "Not Null Flag",
vv.name "Validation Rule"
FROM M0.Entity Pen
JOIN M0.Attribute Pat ON PAt.owner@ = PEn.Id@
left join m0.check_constraint_usage ccu
on Pat.id@ = ccu.owner@
left join m0.validation_rule vr
on vr.id@ = ccu.validation_rule_ref
left join m0.valid_value vv
on vv.owner@ = vr.id@
where Pen.hide_in_logical is null
order by "Entity Name", "Primary Key Flag" desc
            """
extract_defs("Attributes", sql)


#Tidy up sheets
#Remove the 1st sheet
sheet1=wb['Sheet']
wb.remove(sheet1)

#Hide the 1st column and adjust the column widths for all sheets
for sheet in wb.sheetnames:
    ws = wb[sheet]
    #if sheet not in ("Subject Areas","Entities","Relationships"):
    #    ws.delete_cols(1,2)
    ws.column_dimensions['A'].hidden = True
    cols = tuple(ws.columns)
    colno=1
    for col in cols:
        max_col_width=-1
        for row in col:
            if row.value:
                if len(row.value)>max_col_width:
                    max_col_width=len(row.value)
        ws.column_dimensions[get_column_letter(colno)].width = max_col_width+4
        colno=colno+1

    #Add column filters in Entities and Attributes worksheet
    if sheet in ("Entities","Attributes"):
        len_cols = [len(col) for col in ws.columns]
        ws.auto_filter.ref = "B1:B"+str(len_cols[1])
    #Add 3rd column for Attributes worksheet
    if sheet == "Attributes":
        ws.column_dimensions['C'].hidden = True

model_name = fetch_model_name()
wb.save(model_name+" data dictionary.xlsx")
