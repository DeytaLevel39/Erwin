import pyodbc
from openpyxl import load_workbook
wb = load_workbook(filename = 'Definition_Comment_DT_Import_ r9.xlsm')
ws = wb["Attribute Definition"]

from csv import reader
# open file in read mode
with open('Glossary grid 2021-01-11-15-49-17 UTC.csv', 'r',  encoding="utf8") as read_obj:
    # pass the file object to reader() to get the reader object
    csv_reader = reader(read_obj)
    # Iterate over each row in the csv using reader object
    axon_data_elements=[]
    axon_entities = []
    r = 2
    for row in csv_reader:
        if row[1] == "Data Element":
            data_element_info = [row[0],row[2]]
            ws.cell(row=r, column=1).value = row[2]
            ws.cell(row=r, column=2).value = row[0]
            ws.cell(row=r, column=4).value = row[5]
            r+=1
            axon_data_elements.append(data_element_info)
        elif row[1] == "Entity":
            axon_entities.append(row[0])
# Specifying the ODBC driver, server name, database, etc. directly
cnxn = pyodbc.connect('DSN=erwin_r9_current',server='127.0.0.1')
# Create a cursor from the connection
cursor = cnxn.cursor()

def fetch_model_name():
    sql = """
    SELECT 
    TRAN(PEn.owner_path) "Model Name"
    FROM M0.Entity Pen
    """

    cursor.execute(sql)
    model_name = cursor.fetchone()[0]

    return model_name


def extract_defs(type_name, sql):
    cursor.execute(sql)

    # Fetch and display the rows
    resultset = cursor.fetchall()


    for row in resultset:
        if type_name == "Entities":
            erwin_entities.append(row[0])
        else:
            item = [row[0],row[2]]
            erwin_data_elements.append(item)


global erwin_entities
erwin_entities = []

# fetch all of the entities
sql = """
SELECT distinct
      replace(Tran(PEn.Physical_Name),'_',' ')      'Entity Name',
      Tran(PEn.Definition)         'Entity Definition',
   TRAN(PEn.Long_Id) "Entity ID"
FROM
   Entity PEn
where PEn.hide_in_logical is null
order By 1
            """
extract_defs("Entities",sql)

global erwin_data_elements
erwin_data_elements = []

# fetch all of the data elements (attributes)
sql = """
SELECT distinct
TRAN(PAt.Name) "Data Element Name",
TRAN(PAt.Definition) "Data Element Definition",
replace(Tran(PEn.Physical_Name),'_',' ')      'Entity Name'
FROM M0.Entity Pen
JOIN M0.Attribute Pat ON PAt.owner@ = PEn.Id@
left join m0.check_constraint_usage ccu
on Pat.id@ = ccu.owner@
left join m0.validation_rule vr
on vr.id@ = ccu.validation_rule_ref
left join m0.valid_value vv
on vv.owner@ = vr.id@
where Pen.hide_in_logical is null
order by "Data Element Name" desc
            """
extract_defs("Data Elements", sql)

for e in set(axon_entities):
    if e not in erwin_entities:
        print("Erwin is missing entity "+e)

data_elements = [de[0] for de in erwin_data_elements]
for de in axon_data_elements:
    if de[0] not in data_elements:
        print("Erwin is missing data element "+de[1]+"."+de[0])

data_elements = [de[0] for de in axon_data_elements]
for de in erwin_data_elements:
    if de[0] not in data_elements:
        print("Erwin has data element "+de[1]+"."+de[0]+" which isn't in Axon")

wb.save(filename = 'Definition_Comment_DT_Import_ r9.xlsx')


print(axon_data_elements)
print(erwin_data_elements)