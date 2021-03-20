import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

wb_filename=os.environ['WB_FILENAME']
wb = load_workbook(wb_filename)
ws = wb['Entities']
curr_subject_area=""
start=True
for row in range(2, ws.max_row+1):
    subject_area=ws['B' + str(row)].value
    er_diagram = ws['C' + str(row)].value
    name = ws['D' + str(row)].value
    desc = ws['E' + str(row)].value
    if subject_area != curr_subject_area:
        if start!=True:
            n = f.write("</table></html")
            f.close()
        else:
            start=False
        curr_subject_area = subject_area
        header_html = """
        <html>
        <head>
        <title>
        """ + subject_area + """</title>
        </head>
        <body>
        <style type="text/css">
        .tg  {border-collapse:collapse;border-spacing:0;}
        .tg td{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
          overflow:hidden;padding:10px 5px;word-break:normal;}
        .tg th{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
          font-weight:normal;overflow:hidden;padding:10px 5px;word-break:normal;}
        .tg .tg-0lax{text-align:left;vertical-align:top}
        </style>

        """

        f = open(subject_area+".html", "w")
        n = f.write(header_html)
        n = f.write("<H2>"+subject_area+" domain entities</H2><table class=\"tg\"<tr><td><b>E-R Diagram</td><td><b>Entity Name</td><td><b>Entity Description</td></b></tr>")
    n = f.write("<tr><td>{}</td><td>{}</td><td>{}</td></tr>".format(er_diagram, name, desc))
n = f.write("</table></body></html")
f.close()

